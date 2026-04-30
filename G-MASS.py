import webview
import logging
logging.getLogger('pywebview').setLevel(logging.CRITICAL)

import smtplib
import os
import json
import threading
import time
import openpyxl
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from tkinter import filedialog
import tkinter as tk
from cryptography.fernet import Fernet

# ─────────────────────────────────────────────
#  CONFIGURAÇÕES — sem dados pessoais
# ─────────────────────────────────────────────
ASSINATURAS_FILE = "assinaturas.json"
CREDENCIAIS_FILE = "credenciais.dat"
CHAVE_FILE       = "chave.key"
SMTP_HOST        = "smtp.gmail.com"
SMTP_PORTA       = 587
VERSAO           = "1.0.0"


# ─────────────────────────────────────────────
#  CRIPTOGRAFIA — chave única por instalação
# ─────────────────────────────────────────────
def obter_chave():
    """Gera uma chave única na primeira instalação."""
    if not os.path.exists(CHAVE_FILE):
        chave = Fernet.generate_key()
        with open(CHAVE_FILE, "wb") as f:
            f.write(chave)
    with open(CHAVE_FILE, "rb") as f:
        return f.read()

def salvar_credenciais(email, senha):
    try:
        fernet = Fernet(obter_chave())
        dados  = json.dumps({"email": email, "senha": senha})
        with open(CREDENCIAIS_FILE, "wb") as f:
            f.write(fernet.encrypt(dados.encode()))
        return True
    except Exception:
        return False

def carregar_credenciais():
    if not os.path.exists(CREDENCIAIS_FILE):
        return None
    try:
        fernet = Fernet(obter_chave())
        with open(CREDENCIAIS_FILE, "rb") as f:
            dados = fernet.decrypt(f.read())
        return json.loads(dados.decode())
    except Exception:
        # Chave comprometida ou arquivo corrompido — limpa tudo
        limpar_credenciais()
        return None

def limpar_credenciais():
    for f in [CREDENCIAIS_FILE]:
        if os.path.exists(f):
            os.remove(f)


# ─────────────────────────────────────────────
#  ASSINATURAS
# ─────────────────────────────────────────────
def carregar_assinaturas():
    if os.path.exists(ASSINATURAS_FILE):
        with open(ASSINATURAS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def salvar_assinaturas(assinaturas):
    with open(ASSINATURAS_FILE, "w", encoding="utf-8") as f:
        json.dump(assinaturas, f, ensure_ascii=False, indent=2)


# ─────────────────────────────────────────────
#  IMPORTAR EXCEL
# ─────────────────────────────────────────────
def importar_excel(caminho):
    wb   = openpyxl.load_workbook(caminho)
    ws   = wb.active
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        raise ValueError("Planilha vazia.")

    cabecalho = [str(c).strip().lower() if c else "" for c in rows[0]]

    for col in ("nome", "para", "assunto"):
        if col not in cabecalho:
            raise ValueError(f"Coluna obrigatória '{col}' não encontrada.")

    idx_nome    = cabecalho.index("nome")
    idx_para    = cabecalho.index("para")
    idx_assunto = cabecalho.index("assunto")
    idx_cc      = cabecalho.index("cc") if "cc" in cabecalho else None

    contatos = []
    for row in rows[1:]:
        nome    = str(row[idx_nome]).strip()    if row[idx_nome]    else ""
        para    = str(row[idx_para]).strip()    if row[idx_para]    else ""
        assunto = str(row[idx_assunto]).strip() if row[idx_assunto] else ""
        cc      = str(row[idx_cc]).strip()      if idx_cc is not None and row[idx_cc] else ""
        if para and "@" in para:
            contatos.append({"nome": nome, "para": para, "cc": cc, "assunto": assunto})

    if not contatos:
        raise ValueError("Nenhum destinatário válido encontrado.")
    return contatos


# ─────────────────────────────────────────────
#  GERAR MODELO EXCEL
# ─────────────────────────────────────────────
def gerar_modelo_excel(destino):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "destinatarios"

    cabecalho = ["nome", "para", "cc", "assunto"]
    for col, titulo in enumerate(cabecalho, start=1):
        cell = ws.cell(row=1, column=col, value=titulo)
        cell.font      = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill      = openpyxl.styles.PatternFill("solid", fgColor="1a73e8")
        cell.alignment = openpyxl.styles.Alignment(horizontal="center")

    exemplos = [
        ["ADM 1", "adm1@empresa.com", "financeiro@empresa.com", "Assunto do e-mail ADM 1"],
        ["ADM 2", "adm2@empresa.com", "",                       "Assunto do e-mail ADM 2"],
        ["ADM 3", "adm3@empresa.com", "copia@empresa.com",      "Assunto do e-mail ADM 3"],
    ]
    for row_idx, row in enumerate(exemplos, start=2):
        for col_idx, valor in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=valor)

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 30

    wb.save(destino)


# ─────────────────────────────────────────────
#  VALIDAR E-MAIL
# ─────────────────────────────────────────────
def email_valido(email):
    import re
    return bool(re.match(r"^[^\s@]+@[^\s@]+\.[^\s@]+$", email.strip()))


# ─────────────────────────────────────────────
#  ENVIO
# ─────────────────────────────────────────────
def enviar_email(usuario, senha, destinatario, cc, assunto, corpo, caminho_ass):
    try:
        msg = MIMEMultipart("related")
        msg["From"]    = usuario
        msg["To"]      = destinatario
        msg["Subject"] = assunto
        if cc:
            msg["Cc"] = cc

        if caminho_ass and os.path.exists(caminho_ass):
            corpo_html = f"""
            <div style="font-family:Arial,sans-serif;font-size:14px;color:#202124;">
                {corpo.replace(chr(10), '<br>')}
                <br><br>
                <img src="cid:assinatura" style="max-width:600px;">
            </div>"""
        else:
            corpo_html = f"""
            <div style="font-family:Arial,sans-serif;font-size:14px;color:#202124;">
                {corpo.replace(chr(10), '<br>')}
            </div>"""

        msg.attach(MIMEText(corpo_html, "html", "utf-8"))

        if caminho_ass and os.path.exists(caminho_ass):
            with open(caminho_ass, "rb") as f:
                img = MIMEImage(f.read())
            img.add_header("Content-ID", "<assinatura>")
            img.add_header("Content-Disposition", "inline")
            msg.attach(img)

        todos = [destinatario] + ([cc] if cc else [])

        with smtplib.SMTP(SMTP_HOST, SMTP_PORTA) as servidor:
            servidor.ehlo()
            servidor.starttls()
            servidor.ehlo()
            servidor.login(usuario, senha)
            servidor.sendmail(usuario, todos, msg.as_string())

        return True, "Enviado"

    except smtplib.SMTPAuthenticationError:
        return False, "Falha na autenticação"
    except smtplib.SMTPRecipientsRefused:
        return False, "Destinatário recusado"
    except Exception as e:
        return False, str(e)



#  API — exposta ao JavaScript - CUIDADO PARA N VAZAR PELO AMOR
class API:
    def __init__(self):
        self.contatos    = []
        self.assinaturas = carregar_assinaturas()
        self.window      = None
        self.enviando    = False
        self.usuario     = ""
        self.senha       = ""

    # ── Info ──
    def get_versao(self):
        return VERSAO

    # ── Login ──
    def verificar_login_salvo(self):
        """Chamado de forma assíncrona — não bloqueia a UI."""
        cred = carregar_credenciais()
        if cred:
            self.usuario = cred["email"]
            self.senha   = cred["senha"]
            return {"logado": True, "email": self.usuario}
        return {"logado": False}

    def fazer_login(self, email, senha, lembrar):
        email = email.strip().lower()
        senha = senha.strip()

        if not email or not senha:
            return {"ok": False, "erro": "Preencha todos os campos."}
        if not email_valido(email):
            return {"ok": False, "erro": "E-mail inválido."}

        try:
            with smtplib.SMTP(SMTP_HOST, SMTP_PORTA, timeout=10) as servidor:
                servidor.ehlo()
                servidor.starttls()
                servidor.ehlo()
                servidor.login(email, senha)

            self.usuario = email
            self.senha   = senha

            if lembrar:
                salvar_credenciais(email, senha)
            else:
                limpar_credenciais()

            return {"ok": True}

        except smtplib.SMTPAuthenticationError:
            return {"ok": False, "erro": "E-mail ou senha de app incorretos."}
        except TimeoutError:
            return {"ok": False, "erro": "Tempo de conexão esgotado. Verifique sua internet."}
        except Exception as e:
            return {"ok": False, "erro": f"Erro de conexão: {str(e)}"}

    def fazer_logout(self):
        self.usuario = ""
        self.senha   = ""
        self.contatos = []
        limpar_credenciais()
        return {"ok": True}

    def abrir_guia_senha_app(self):
        import webbrowser
        webbrowser.open("https://myaccount.google.com/apppasswords")
        return {"ok": True}

    # ── Modelo Excel ──
    def baixar_modelo(self):
        root = tk.Tk()
        root.withdraw()
        destino = filedialog.asksaveasfilename(
            title            = "Salvar modelo como",
            defaultextension = ".xlsx",
            initialfile      = "modelo_destinatarios.xlsx",
            filetypes        = [("Excel", "*.xlsx")])
        root.destroy()
        if not destino:
            return {"ok": False, "erro": "Cancelado."}
        try:
            gerar_modelo_excel(destino)
            return {"ok": True, "arquivo": os.path.basename(destino)}
        except Exception as e:
            return {"ok": False, "erro": str(e)}

    # ── Assinaturas ──
    def get_assinaturas(self):
        return list(self.assinaturas.keys())

    def adicionar_assinatura(self, nome):
        nome = nome.strip()
        if not nome:
            return {"ok": False, "erro": "Nome inválido."}
        root = tk.Tk()
        root.withdraw()
        caminho = filedialog.askopenfilename(
            title     = "Selecione a imagem",
            filetypes = [("Imagens", "*.jpg *.jpeg *.png")])
        root.destroy()
        if not caminho:
            return {"ok": False, "erro": "Nenhuma imagem selecionada."}
        self.assinaturas[nome] = caminho
        salvar_assinaturas(self.assinaturas)
        return {"ok": True, "nomes": list(self.assinaturas.keys())}

    def remover_assinatura(self, nome):
        if nome in self.assinaturas:
            del self.assinaturas[nome]
            salvar_assinaturas(self.assinaturas)
        return {"ok": True, "nomes": list(self.assinaturas.keys())}

    # ── Excel ──
    def selecionar_excel(self):
        root = tk.Tk()
        root.withdraw()
        caminho = filedialog.askopenfilename(
            title     = "Selecione a planilha",
            filetypes = [("Excel", "*.xlsx")])
        root.destroy()
        if not caminho:
            return {"ok": False, "erro": "Nenhum arquivo selecionado."}
        try:
            self.contatos = importar_excel(caminho)
            return {
                "ok":       True,
                "arquivo":  os.path.basename(caminho),
                "total":    len(self.contatos),
                "contatos": self.contatos
            }
        except Exception as e:
            return {"ok": False, "erro": str(e)}

    # ── Envio ──
    def iniciar_envio(self, corpo, nome_assinatura, delay):
        if self.enviando:
            return {"ok": False, "erro": "Envio já em andamento."}
        if not self.contatos:
            return {"ok": False, "erro": "Nenhum contato carregado."}
        if not corpo.strip():
            return {"ok": False, "erro": "Mensagem vazia."}
        if not self.usuario or not self.senha:
            return {"ok": False, "erro": "Usuário não autenticado."}

        self.enviando = True
        threading.Thread(
            target=self._enviar_lote,
            args=(corpo, nome_assinatura, max(0, int(delay))),
            daemon=True
        ).start()
        return {"ok": True}

    def _enviar_lote(self, corpo, nome_assinatura, delay):
        caminho_ass = self.assinaturas.get(nome_assinatura, "")
        total   = len(self.contatos)
        sucesso = 0
        erros   = 0

        for i, contato in enumerate(self.contatos, start=1):
            self.window.evaluate_js(
                f"atualizarStatus({i-1}, 'enviando', 'Enviando...')")
            self.window.evaluate_js(
                f"atualizarProgresso({i}, {total})")

            ok, msg = enviar_email(
                self.usuario, self.senha,
                contato["para"], contato["cc"],
                contato["assunto"], corpo, caminho_ass)

            if ok:
                sucesso += 1
                self.window.evaluate_js(
                    f"atualizarStatus({i-1}, 'ok', '✓ Enviado')")
            else:
                erros += 1
                msg_e = msg.replace("'", "\\'").replace("\n", " ")
                self.window.evaluate_js(
                    f"atualizarStatus({i-1}, 'erro', '✗ {msg_e}')")

            if i < total and delay > 0:
                for seg in range(delay):
                    restante = delay - seg
                    self.window.evaluate_js(
                        f"atualizarDelay({restante}, {delay})")
                    time.sleep(1)
                self.window.evaluate_js("limparDelay()")

        self.enviando = False
        self.window.evaluate_js(f"finalizarEnvio({sucesso}, {erros})")


#  INICIAR
if __name__ == "__main__":
    api    = API()
    window = webview.create_window(
        title            = "G-MASS",
        url              = "index.html",
        js_api           = api,
        width            = 900,
        height           = 700,
        resizable        = True,
        min_size         = (800, 600),
        background_color = "#0f0f1a"
    )
    api.window = window
    webview.start()